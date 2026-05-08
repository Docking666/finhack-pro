"""
Excel导出器

使用openpyxl库生成多工作表的Excel文件。
工作表：摘要、权益曲线、交易记录、持仓记录。
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from loguru import logger

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Fill,
    PatternFill,
    Border,
    Side,
    Alignment,
    NamedStyle,
)
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel导出器类
    
    用于将回测结果导出为格式化的Excel文件。
    """
    
    # 品牌颜色
    BRAND_COLOR = "3B82F6"  # 蓝色
    SUCCESS_COLOR = "22C55E"  # 绿色
    DANGER_COLOR = "EF4444"   # 红色
    WARNING_COLOR = "F59E0B"  # 橙色
    HEADER_COLOR = "1E293B"   # 深蓝灰
    ALT_ROW_COLOR = "F8FAFC"  # 交替行颜色
    
    def __init__(self):
        """初始化Excel导出器"""
        self._create_styles()
    
    def _create_styles(self):
        """创建自定义样式"""
        # 边框样式
        self.thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )
        
        # 填充样式
        self.header_fill = PatternFill(
            start_color=self.HEADER_COLOR,
            end_color=self.HEADER_COLOR,
            fill_type='solid'
        )
        self.brand_fill = PatternFill(
            start_color=self.BRAND_COLOR,
            end_color=self.BRAND_COLOR,
            fill_type='solid'
        )
        self.alt_row_fill = PatternFill(
            start_color=self.ALT_ROW_COLOR,
            end_color=self.ALT_ROW_COLOR,
            fill_type='solid'
        )
        self.success_fill = PatternFill(
            start_color=self.SUCCESS_COLOR,
            end_color=self.SUCCESS_COLOR,
            fill_type='solid'
        )
        self.danger_fill = PatternFill(
            start_color=self.DANGER_COLOR,
            end_color=self.DANGER_COLOR,
            fill_type='solid'
        )
        
        # 字体样式
        self.title_font = Font(name='Arial', size=18, bold=True, color=self.HEADER_COLOR)
        self.header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.body_font = Font(name='Arial', size=10, color='374151')
        self.number_font = Font(name='Consolas', size=10, color='374151')
        self.success_font = Font(name='Arial', size=10, color=self.SUCCESS_COLOR)
        self.danger_font = Font(name='Arial', size=10, color=self.DANGER_COLOR)
    
    def export_backtest_report(
        self,
        result: Dict[str, Any],
        output_path: str,
        params: Optional[Dict[str, Any]] = None,
    ) -> str:
        """导出回测报告为Excel
        
        Args:
            result: 回测结果数据
            output_path: 输出文件路径
            params: 回测参数（可选）
        
        Returns:
            生成的Excel文件路径
        """
        wb = Workbook()
        
        # 删除默认工作表
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 1. 创建摘要工作表
        self._create_summary_sheet(wb, result, params)
        
        # 2. 创建权益曲线工作表
        if result.get("equity_curve"):
            self._create_equity_sheet(wb, result)
        
        # 3. 创建交易记录工作表
        if result.get("trades"):
            self._create_trades_sheet(wb, result)
        
        # 4. 创建基准对比工作表
        if result.get("benchmark_curve"):
            self._create_benchmark_sheet(wb, result)
        
        # 保存文件
        wb.save(output_path)
        logger.info(f"Excel报告已生成: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        result: Dict[str, Any],
        params: Optional[Dict[str, Any]] = None,
    ):
        """创建摘要工作表"""
        ws = wb.create_sheet("摘要", 0)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        current_row = 1
        
        # 标题
        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws['B' + str(current_row)] = "FinHack Pro 回测报告"
        ws['B' + str(current_row)].font = self.title_font
        ws['B' + str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2
        
        # 基本信息
        ws['B' + str(current_row)] = "基本信息"
        ws['B' + str(current_row)].font = Font(name='Arial', size=12, bold=True, color=self.BRAND_COLOR)
        current_row += 1
        
        params = params or {}
        basic_info = [
            ("策略名称", params.get("strategy", "Dual Thrust")),
            ("标的代码", params.get("symbols", "N/A")),
            ("回测区间", f"{params.get('start_date', '')} ~ {params.get('end_date', '')}"),
            ("初始资金", f"¥{params.get('initial_capital', 1000000):,.0f}"),
            ("基准指数", params.get("benchmark", "000300.SH")),
            ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        
        for label, value in basic_info:
            ws['B' + str(current_row)] = label
            ws['B' + str(current_row)].font = self.body_font
            ws['B' + str(current_row)].alignment = Alignment(horizontal='right')
            
            ws['C' + str(current_row)] = value
            ws['C' + str(current_row)].font = Font(name='Arial', size=10, bold=True)
            
            current_row += 1
        
        current_row += 2
        
        # 关键指标
        ws['B' + str(current_row)] = "关键指标"
        ws['B' + str(current_row)].font = Font(name='Arial', size=12, bold=True, color=self.BRAND_COLOR)
        current_row += 1
        
        # 指标表头
        headers = ["指标名称", "数值", "说明"]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        current_row += 1
        
        # 指标数据
        metrics = result.get("metrics", {})
        metrics_data = [
            ("总收益率", f"{metrics.get('total_return', 0):.2f}%", "整个回测期间的累计收益", metrics.get('total_return', 0) >= 0),
            ("年化收益率", f"{metrics.get('annual_return', 0):.2f}%", "折算为年度的收益率", metrics.get('annual_return', 0) >= 0),
            ("夏普比率", f"{metrics.get('sharpe_ratio', 0):.2f}", "风险调整后收益指标", metrics.get('sharpe_ratio', 0) >= 1),
            ("Sortino比率", f"{metrics.get('sortino_ratio', 0):.2f}", "下行风险调整后收益", metrics.get('sortino_ratio', 0) >= 1),
            ("最大回撤", f"{metrics.get('max_drawdown', 0):.2f}%", "最大峰值到谷值的跌幅", metrics.get('max_drawdown', 0) <= 20),
            ("胜率", f"{metrics.get('win_rate', 0):.2f}%", "盈利交易占比", metrics.get('win_rate', 0) >= 50),
            ("盈亏比", f"{metrics.get('profit_loss_ratio', 0):.2f}", "平均盈利/平均亏损", metrics.get('profit_loss_ratio', 0) >= 1),
            ("交易次数", f"{metrics.get('total_trades', 0)}", "总交易次数", None),
            ("最终权益", f"¥{metrics.get('final_equity', 0):,.2f}", "期末账户权益", None),
        ]
        
        for i, (name, value, desc, is_good) in enumerate(metrics_data):
            # 交替行背景
            if i % 2 == 1:
                for col in range(2, 5):
                    ws.cell(row=current_row, column=col).fill = self.alt_row_fill
            
            ws.cell(row=current_row, column=2, value=name).font = self.body_font
            ws.cell(row=current_row, column=2).border = self.thin_border
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='left')
            
            value_cell = ws.cell(row=current_row, column=3, value=value)
            value_cell.font = self.number_font
            value_cell.border = self.thin_border
            value_cell.alignment = Alignment(horizontal='right')
            
            # 根据指标设置颜色
            if is_good is not None:
                if is_good:
                    value_cell.font = self.success_font
                else:
                    value_cell.font = self.danger_font
            
            ws.cell(row=current_row, column=4, value=desc).font = Font(name='Arial', size=9, color='6B7280')
            ws.cell(row=current_row, column=4).border = self.thin_border
            
            current_row += 1
        
        # 设置行高
        for row in range(1, current_row + 1):
            ws.row_dimensions[row].height = 22
    
    def _create_equity_sheet(self, wb: Workbook, result: Dict[str, Any]):
        """创建权益曲线工作表"""
        ws = wb.create_sheet("权益曲线", 1)
        
        equity_curve = result.get("equity_curve", [])
        
        # 设置列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        
        # 标题
        ws['B1'] = "权益曲线数据"
        ws['B1'].font = Font(name='Arial', size=14, bold=True, color=self.BRAND_COLOR)
        
        # 表头
        headers = ["日期", "权益", "收益率", "回撤"]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 数据
        initial_equity = equity_curve[0].get("equity", equity_curve[0].get("value", 1000000)) if equity_curve else 1000000
        peak = initial_equity
        
        for i, point in enumerate(equity_curve):
            row = i + 4
            equity = point.get("equity", point.get("value", 0))
            return_pct = (equity - initial_equity) / initial_equity * 100 if initial_equity else 0
            
            # 计算回撤
            if equity > peak:
                peak = equity
            drawdown = (peak - equity) / peak * 100 if peak else 0
            
            # 日期
            ws.cell(row=row, column=2, value=point.get("date", "")).font = self.body_font
            ws.cell(row=row, column=2).border = self.thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            # 权益
            equity_cell = ws.cell(row=row, column=3, value=equity)
            equity_cell.font = self.number_font
            equity_cell.border = self.thin_border
            equity_cell.alignment = Alignment(horizontal='right')
            equity_cell.number_format = '#,##0.00'
            
            # 收益率
            return_cell = ws.cell(row=row, column=4, value=return_pct / 100)
            return_cell.font = self.success_font if return_pct >= 0 else self.danger_font
            return_cell.border = self.thin_border
            return_cell.alignment = Alignment(horizontal='right')
            return_cell.number_format = '0.00%'
            
            # 回撤
            dd_cell = ws.cell(row=row, column=5, value=drawdown / 100)
            dd_cell.font = self.danger_font if drawdown > 0 else self.body_font
            dd_cell.border = self.thin_border
            dd_cell.alignment = Alignment(horizontal='right')
            dd_cell.number_format = '0.00%'
            
            # 交替行背景
            if i % 2 == 1:
                for col in range(2, 6):
                    ws.cell(row=row, column=col).fill = self.alt_row_fill
        
        # 创建权益曲线图表
        if len(equity_curve) > 1:
            chart = LineChart()
            chart.title = "权益曲线"
            chart.style = 10
            chart.y_axis.title = "权益"
            chart.x_axis.title = "日期"
            chart.width = 20
            chart.height = 12
            
            # 数据范围
            data = Reference(ws, min_col=3, min_row=3, max_col=3, max_row=len(equity_curve) + 3)
            chart.add_data(data, titles_from_data=True)
            
            # 设置线条颜色
            chart.series[0].graphicalProperties.line.solidFill = self.BRAND_COLOR
            chart.series[0].graphicalProperties.line.width = 25000  # 2.5pt
            
            ws.add_chart(chart, "G3")
    
    def _create_trades_sheet(self, wb: Workbook, result: Dict[str, Any]):
        """创建交易记录工作表"""
        ws = wb.create_sheet("交易记录", 2)
        
        trades = result.get("trades", [])
        
        # 设置列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 20
        
        # 标题
        ws['B1'] = "交易记录"
        ws['B1'].font = Font(name='Arial', size=14, bold=True, color=self.BRAND_COLOR)
        
        # 统计信息
        ws['B2'] = f"共 {len(trades)} 笔交易"
        ws['B2'].font = Font(name='Arial', size=10, color='6B7280')
        
        # 表头
        headers = ["日期", "标的", "方向", "价格", "数量", "金额", "盈亏", "原因"]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 数据
        for i, trade in enumerate(trades):
            row = i + 5
            
            direction = trade.get("direction", "buy")
            direction_text = "买入" if direction == "buy" else "卖出"
            price = trade.get("price", 0)
            volume = trade.get("volume", 0)
            amount = price * volume
            pnl = trade.get("pnl", 0)
            
            # 日期
            ws.cell(row=row, column=2, value=trade.get("date", "")).font = self.body_font
            ws.cell(row=row, column=2).border = self.thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            # 标的
            ws.cell(row=row, column=3, value=trade.get("symbol", "")).font = self.number_font
            ws.cell(row=row, column=3).border = self.thin_border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            # 方向
            dir_cell = ws.cell(row=row, column=4, value=direction_text)
            dir_cell.font = self.success_font if direction == "buy" else self.danger_font
            dir_cell.border = self.thin_border
            dir_cell.alignment = Alignment(horizontal='center')
            
            # 价格
            price_cell = ws.cell(row=row, column=5, value=price)
            price_cell.font = self.number_font
            price_cell.border = self.thin_border
            price_cell.alignment = Alignment(horizontal='right')
            price_cell.number_format = '#,##0.00'
            
            # 数量
            ws.cell(row=row, column=6, value=volume).font = self.number_font
            ws.cell(row=row, column=6).border = self.thin_border
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=6).number_format = '#,##0'
            
            # 金额
            amount_cell = ws.cell(row=row, column=7, value=amount)
            amount_cell.font = self.number_font
            amount_cell.border = self.thin_border
            amount_cell.alignment = Alignment(horizontal='right')
            amount_cell.number_format = '#,##0.00'
            
            # 盈亏
            pnl_cell = ws.cell(row=row, column=8, value=pnl if pnl else 0)
            pnl_cell.font = self.success_font if pnl > 0 else (self.danger_font if pnl < 0 else self.number_font)
            pnl_cell.border = self.thin_border
            pnl_cell.alignment = Alignment(horizontal='right')
            pnl_cell.number_format = '#,##0.00'
            
            # 原因
            ws.cell(row=row, column=9, value=trade.get("reason", "")).font = Font(name='Arial', size=9, color='6B7280')
            ws.cell(row=row, column=9).border = self.thin_border
            
            # 交替行背景
            if i % 2 == 1:
                for col in range(2, 10):
                    ws.cell(row=row, column=col).fill = self.alt_row_fill
        
        # 冻结首行
        ws.freeze_panes = 'B5'
    
    def _create_benchmark_sheet(self, wb: Workbook, result: Dict[str, Any]):
        """创建基准对比工作表"""
        ws = wb.create_sheet("基准对比", 3)
        
        benchmark_curve = result.get("benchmark_curve", [])
        equity_curve = result.get("equity_curve", [])
        
        if not benchmark_curve:
            return
        
        # 设置列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # 标题
        ws['B1'] = "基准对比"
        ws['B1'].font = Font(name='Arial', size=14, bold=True, color=self.BRAND_COLOR)
        
        # 表头
        headers = ["日期", "策略权益", "基准权益", "超额收益"]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 数据
        for i, point in enumerate(benchmark_curve):
            row = i + 4
            
            equity_point = equity_curve[i] if i < len(equity_curve) else {}
            strategy_equity = equity_point.get("equity", equity_point.get("value", 0))
            benchmark_equity = point.get("equity", point.get("value", 0))
            
            # 计算超额收益
            if i > 0 and benchmark_curve[i-1].get("equity", benchmark_curve[i-1].get("value", 0)) > 0:
                prev_benchmark = benchmark_curve[i-1].get("equity", benchmark_curve[i-1].get("value", 0))
                prev_strategy = equity_curve[i-1].get("equity", equity_curve[i-1].get("value", 0)) if i-1 < len(equity_curve) else prev_benchmark
                
                strategy_return = (strategy_equity - prev_strategy) / prev_strategy if prev_strategy else 0
                benchmark_return = (benchmark_equity - prev_benchmark) / prev_benchmark if prev_benchmark else 0
                excess_return = strategy_return - benchmark_return
            else:
                excess_return = 0
            
            # 日期
            ws.cell(row=row, column=2, value=point.get("date", "")).font = self.body_font
            ws.cell(row=row, column=2).border = self.thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            # 策略权益
            ws.cell(row=row, column=3, value=strategy_equity).font = self.number_font
            ws.cell(row=row, column=3).border = self.thin_border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            
            # 基准权益
            ws.cell(row=row, column=4, value=benchmark_equity).font = self.number_font
            ws.cell(row=row, column=4).border = self.thin_border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            
            # 超额收益
            excess_cell = ws.cell(row=row, column=5, value=excess_return)
            excess_cell.font = self.success_font if excess_return >= 0 else self.danger_font
            excess_cell.border = self.thin_border
            excess_cell.alignment = Alignment(horizontal='right')
            excess_cell.number_format = '0.00%'
            
            # 交替行背景
            if i % 2 == 1:
                for col in range(2, 6):
                    ws.cell(row=row, column=col).fill = self.alt_row_fill
    
    def export_to_bytes(
        self,
        result: Dict[str, Any],
        params: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """导出回测报告为字节流
        
        Args:
            result: 回测结果数据
            params: 回测参数（可选）
        
        Returns:
            Excel文件的字节流
        """
        buffer = io.BytesIO()
        
        wb = Workbook()
        
        # 删除默认工作表
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 创建各个工作表
        self._create_summary_sheet(wb, result, params)
        
        if result.get("equity_curve"):
            self._create_equity_sheet(wb, result)
        
        if result.get("trades"):
            self._create_trades_sheet(wb, result)
        
        if result.get("benchmark_curve"):
            self._create_benchmark_sheet(wb, result)
        
        # 保存到字节流
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
